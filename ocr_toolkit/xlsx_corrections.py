"""完成済み XLSX をシート単位で LLM 校正辞書で後校正するユーティリティ。

想定ワークフロー:

    完成 XLSX → load_sheet_rows → apply_corrections → openpyxl で書き戻し

既存の `ocr_toolkit.corrections.apply_corrections` は rows (dict リスト) ベースで
DB 投入を想定しているが、本モジュールは「成果物として XLSX を保持しつつ誤字を
差し替える」後校正ユースケースを担当する。

校正辞書ファイル形式 (既存 CORRECTIONS dict を踏襲):

    # corrections/llm_corrections_sheet1_001.py
    CORRECTIONS = {
        ("Destiny", 1): {"Body": "...修正後..."},
        ("Soul", 11):   {"Title1": "...", "Body": "..."},
    }

キーは key_fields で指定した列値のタプル。値は dict[列名, 修正後テキスト]。

本モジュールは openpyxl に依存する (preview.py と同じく optional import)。
"""
from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:  # pragma: no cover
    raise ImportError(
        "xlsx_corrections.py は openpyxl に依存します。`pip install openpyxl` してください。"
    ) from e

from .book_config import BookConfig, SheetConfig, load_book_config
from .corrections import (
    CorrectionError,
    Corrections,
    apply_corrections,
    diff_summary,
    load_merged,
    validate,
)

# preview.py の FILL_CHANGED と揃える (薄黄)
FILL_CHANGED = PatternFill("solid", fgColor="FFFDE7")


# ------------------------------------------------------------
# シート → rows
# ------------------------------------------------------------

def load_sheet_rows(
    xlsx_path: str | Path,
    sheet_name: str,
    *,
    header_row: int = 1,
    data_only: bool = True,
) -> tuple[list[str], list[dict[str, Any]]]:
    """XLSX の 1 シートを (headers, rows) に変換する。

    Args:
        xlsx_path: XLSX ファイルパス。
        sheet_name: 対象シート名。
        header_row: ヘッダ行番号 (1-origin、デフォルト 1 行目)。
        data_only: True で数式の評価済み値を読む (デフォルト True)。

    Returns:
        (headers, rows)
        - headers: 列名のリスト (None は "" に変換)
        - rows: dict[列名, 値] のリスト。空のヘッダ列は含めない。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=data_only)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート '{sheet_name}' が見つかりません: {wb.sheetnames}")
    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < header_row:
        return [], []

    raw_headers = all_rows[header_row - 1]
    headers = [("" if h is None else str(h)) for h in raw_headers]

    rows: list[dict[str, Any]] = []
    for raw in all_rows[header_row:]:
        row: dict[str, Any] = {}
        for h, v in zip(headers, raw):
            if not h:
                continue
            row[h] = v
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)
    return headers, rows


def _make_key_fn(key_fields: Sequence[str]):
    """key_fields のタプルから CorrectionKey を作る関数を返す。

    値は文字列化を試みるが、int キーの書籍もあるため元の型を優先する。
    """
    def key_fn(row: Mapping[str, Any]):
        values = []
        for f in key_fields:
            v = row.get(f)
            values.append(v)
        return tuple(values) if len(values) > 1 else values[0]
    return key_fn


# ------------------------------------------------------------
# XLSX 後校正
# ------------------------------------------------------------

def apply_to_xlsx(
    xlsx_path: str | Path,
    sheet_name: str,
    corrections: Corrections,
    *,
    key_fields: Sequence[str],
    out_path: str | Path | None = None,
    header_row: int = 1,
    highlight: bool = True,
    dry_run: bool = False,
    allowed_fields: Iterable[str] | None = None,
) -> dict[str, Any]:
    """XLSX の指定シートに校正辞書を適用して保存する。

    Args:
        xlsx_path: 入力 XLSX。
        sheet_name: 対象シート。
        corrections: load_merged() で読んだ校正辞書。
        key_fields: 行キーとする列名のタプル (例: ("Type", "Number"))。
        out_path: 出力先。None なら xlsx_path を上書き。
        header_row: ヘッダ行 (1-origin)。
        highlight: True で変更セルに薄黄塗り。
        dry_run: True なら保存せず件数サマリのみ返す。
        allowed_fields: 指定時、CORRECTIONS が含むフィールドを制限 (バリデーション)。
            None なら対象シートのヘッダ列を自動設定。

    Returns:
        {
            "rows": 行数,
            "matched": キーがマッチした行数,
            "applied": 実適用行数,
            "cells_changed": 変更セル数,
            "fields": {列名: 変更件数},
            "unmatched_keys": [corrections にあったが行には無いキー],
        }
    """
    xlsx_path = Path(xlsx_path)
    out_path = Path(out_path) if out_path else xlsx_path

    # 1) 事前バリデーション
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート '{sheet_name}' が見つかりません: {wb.sheetnames}")
    ws: Worksheet = wb[sheet_name]

    raw_headers = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [("" if h is None else str(h)) for h in raw_headers]
    header_set = {h for h in headers if h}

    validate(
        corrections,
        allowed_fields=set(allowed_fields) if allowed_fields is not None else header_set,
    )

    for f in key_fields:
        if f not in header_set:
            raise CorrectionError(f"key_fields '{f}' がシートヘッダに存在しません: {sorted(header_set)}")

    col_index: dict[str, int] = {h: i + 1 for i, h in enumerate(headers) if h}

    # 2) rows 化 + apply_corrections で「変更後 dict リスト」を得る
    _, rows = load_sheet_rows(xlsx_path, sheet_name, header_row=header_row, data_only=False)
    key_fn = _make_key_fn(key_fields)

    # 事前に元の値を保持 (diff 集計用)
    before_by_key: dict[Any, dict[str, Any]] = {}
    for row in rows:
        k = key_fn(row)
        before_by_key[k] = {f: row.get(f) for f in header_set}

    result = apply_corrections(rows, corrections, key_fn=key_fn, dry_run=dry_run)

    # 3) どの key がマッチしなかったか (辞書側には有るが行に無い)
    matched_keys = {key_fn(r) for r in rows if key_fn(r) in corrections}
    unmatched_keys = [k for k in corrections if k not in matched_keys]

    fields_changed = diff_summary(before_by_key, corrections)
    cells_changed = sum(fields_changed.values())

    summary: dict[str, Any] = {
        "rows": result.get("total", 0),
        "matched": result.get("matched", 0),
        "applied": result.get("applied", 0),
        "cells_changed": cells_changed,
        "fields": fields_changed,
        "unmatched_keys": unmatched_keys,
    }

    if dry_run:
        return summary

    # 4) openpyxl で該当セルを書き換え + ハイライト
    for excel_row_idx, row in enumerate(rows, start=header_row + 1):
        k = key_fn(row)
        patch = corrections.get(k)
        if not patch:
            continue
        for field, new_val in patch.items():
            col = col_index.get(field)
            if col is None:
                continue
            cell = ws.cell(row=excel_row_idx, column=col)
            old_val = cell.value
            if (old_val or "") == (new_val or ""):
                continue
            cell.value = new_val
            if highlight:
                cell.fill = FILL_CHANGED

    wb.save(out_path)
    return summary


# ------------------------------------------------------------
# 比較 XLSX プレビュー
# ------------------------------------------------------------

def preview_xlsx_corrections(
    xlsx_path: str | Path,
    sheet_name: str,
    corrections: Corrections,
    *,
    key_fields: Sequence[str],
    output: str | Path,
    fields: Sequence[str] | None = None,
    header_row: int = 1,
) -> dict[str, int]:
    """校正前/後を並べた比較 XLSX を生成する (preview.write_comparison の薄いラッパ)。

    Args:
        xlsx_path: 入力 XLSX。
        sheet_name: 対象シート。
        corrections: 校正辞書。
        key_fields: 行キー列名。
        output: 出力 XLSX パス。
        fields: 比較する列名。None なら corrections に登場する全列。
        header_row: ヘッダ行。

    Returns:
        {"rows":N, "changed":M, "warnings":W}
    """
    from .preview import write_comparison

    headers, rows = load_sheet_rows(xlsx_path, sheet_name, header_row=header_row)
    header_set = {h for h in headers if h}

    if fields is None:
        collected: list[str] = []
        seen: set[str] = set()
        for patch in corrections.values():
            for f in patch:
                if f in header_set and f not in seen:
                    collected.append(f)
                    seen.add(f)
        fields = collected

    key_fn = _make_key_fn(key_fields)

    comparison_rows: list[dict[str, Any]] = []
    for row in rows:
        k = key_fn(row)
        patch = corrections.get(k, {})
        before = {f: row.get(f) for f in fields}
        after = {f: patch.get(f, row.get(f)) for f in fields}
        label_parts = [str(row.get(lf, "")) for lf in key_fields]
        comparison_rows.append({
            "label": "/".join(p for p in label_parts if p),
            "_before": before,
            "_after": after,
        })

    return write_comparison(
        output=Path(output),
        rows=comparison_rows,
        fields=fields,
        label_fields=("label",),
        sheet_title=f"{sheet_name}_diff",
    )


# ------------------------------------------------------------
# 書籍一括校正 (sheets.yaml ベース)
# ------------------------------------------------------------

def apply_book(
    config: BookConfig | str | Path,
    *,
    dry_run: bool = False,
    preview_dir: str | Path | None = None,
    out_path: str | Path | None = None,
    highlight: bool = True,
) -> dict[str, Any]:
    """書籍の sheets.yaml に基づいて enabled な全シートに校正を適用する。

    Args:
        config: BookConfig か sheets.yaml のパス。
        dry_run: True なら保存せず件数のみ集計。
        preview_dir: 指定時、シート毎に <preview_dir>/<sheet>_diff.xlsx を出力。
        out_path: 出力 XLSX。None なら BookConfig.xlsx を上書き。
            dry_run=True のときも入力 XLSX は一切触らない。
        highlight: 変更セルをハイライトするか。

    Returns:
        {
            "<sheet_name>": {<apply_to_xlsx summary>},
            ...,
            "_total": {"sheets_processed": N, "cells_changed": M, "matched": K, ...},
        }

    Notes:
        校正は in-memory で順次積み上がり、全シート処理後に 1 回だけ保存する
        (openpyxl は複数シートを同じ wb に保持するため効率的)。
    """
    if not isinstance(config, BookConfig):
        config = load_book_config(config)

    src_xlsx = config.xlsx
    if not src_xlsx.exists():
        raise FileNotFoundError(f"XLSX が見つかりません: {src_xlsx}")

    dst_xlsx = Path(out_path) if out_path else src_xlsx

    preview_dir = Path(preview_dir) if preview_dir else None
    if preview_dir:
        preview_dir.mkdir(parents=True, exist_ok=True)

    # 書籍 XLSX を 1 回だけ開いて、シート毎に差し替える
    wb = openpyxl.load_workbook(src_xlsx)

    reports: dict[str, Any] = {}
    total = {
        "sheets_processed": 0,
        "sheets_skipped": 0,
        "rows": 0,
        "matched": 0,
        "applied": 0,
        "cells_changed": 0,
    }

    for sheet in config.enabled_sheets():
        if sheet.name not in wb.sheetnames:
            reports[sheet.name] = {"error": "シートが存在しません", "skipped": True}
            total["sheets_skipped"] += 1
            continue

        corr_dir = config.sheet_corrections_dir(sheet)
        if not corr_dir.exists():
            reports[sheet.name] = {"corrections_dir": str(corr_dir), "loaded_files": 0, "skipped": True}
            total["sheets_skipped"] += 1
            continue

        corr_paths = sorted(corr_dir.glob(sheet.glob))
        if not corr_paths:
            reports[sheet.name] = {"corrections_dir": str(corr_dir), "loaded_files": 0, "skipped": True}
            total["sheets_skipped"] += 1
            continue

        corrections = load_merged(corr_paths)
        if not corrections:
            reports[sheet.name] = {"corrections_dir": str(corr_dir), "loaded_files": len(corr_paths), "entries": 0, "skipped": True}
            total["sheets_skipped"] += 1
            continue

        # preview XLSX (任意)
        if preview_dir is not None:
            preview_xlsx_corrections(
                src_xlsx, sheet.name, corrections,
                key_fields=tuple(sheet.key_cols),
                output=preview_dir / f"{sheet.name}_diff.xlsx",
            )

        # wb 上のシートに対して in-memory で apply (src_xlsx を再オープンしない)
        summary = _apply_in_workbook(
            wb, sheet.name, corrections,
            key_fields=tuple(sheet.key_cols),
            highlight=highlight,
            dry_run=dry_run,
        )
        summary["loaded_files"] = len(corr_paths)
        summary["entries"] = len(corrections)
        reports[sheet.name] = summary

        total["sheets_processed"] += 1
        for k in ("rows", "matched", "applied", "cells_changed"):
            total[k] += summary.get(k, 0)

    if not dry_run:
        wb.save(dst_xlsx)

    reports["_total"] = total
    return reports


def _apply_in_workbook(
    wb: "openpyxl.Workbook",
    sheet_name: str,
    corrections: Corrections,
    *,
    key_fields: Sequence[str],
    header_row: int = 1,
    highlight: bool = True,
    dry_run: bool = False,
) -> dict[str, Any]:
    """既に開いている workbook に対して 1 シート分の校正を適用する内部関数。

    apply_to_xlsx とロジックは同じだが、wb を引数で受けるので複数シート処理時に
    load_workbook/save を繰り返さずに済む。
    """
    ws: Worksheet = wb[sheet_name]
    raw_headers = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    headers = [("" if h is None else str(h)) for h in raw_headers]
    header_set = {h for h in headers if h}

    validate(corrections, allowed_fields=header_set)
    for f in key_fields:
        if f not in header_set:
            raise CorrectionError(f"{sheet_name}: key_fields '{f}' がシートヘッダに存在しません: {sorted(header_set)}")

    col_index = {h: i + 1 for i, h in enumerate(headers) if h}
    key_fn = _make_key_fn(key_fields)

    rows: list[dict[str, Any]] = []
    for raw in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row = {h: v for h, v in zip(headers, raw) if h}
        if any(v not in (None, "") for v in row.values()):
            rows.append(row)

    before_by_key: dict[Any, dict[str, Any]] = {
        key_fn(r): {f: r.get(f) for f in header_set} for r in rows
    }
    result = apply_corrections(rows, corrections, key_fn=key_fn, dry_run=dry_run)
    matched_keys = {key_fn(r) for r in rows if key_fn(r) in corrections}
    unmatched_keys = [k for k in corrections if k not in matched_keys]
    fields_changed = diff_summary(before_by_key, corrections)
    cells_changed = sum(fields_changed.values())

    if not dry_run:
        for excel_row_idx, row in enumerate(rows, start=header_row + 1):
            k = key_fn(row)
            patch = corrections.get(k)
            if not patch:
                continue
            for field, new_val in patch.items():
                col = col_index.get(field)
                if col is None:
                    continue
                cell = ws.cell(row=excel_row_idx, column=col)
                if (cell.value or "") == (new_val or ""):
                    continue
                cell.value = new_val
                if highlight:
                    cell.fill = FILL_CHANGED

    return {
        "rows": result.get("total", 0),
        "matched": result.get("matched", 0),
        "applied": result.get("applied", 0),
        "cells_changed": cells_changed,
        "fields": fields_changed,
        "unmatched_keys": unmatched_keys,
    }


__all__ = [
    "load_sheet_rows",
    "apply_to_xlsx",
    "preview_xlsx_corrections",
    "apply_book",
    "FILL_CHANGED",
]
