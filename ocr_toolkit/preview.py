"""OCR → 正規化 → LLM 校正後 を並べた比較 XLSX を生成するユーティリティ。

用途:
- 校正前後の差分を目視確認する
- どのセルが変更されたかをハイライトで把握する
- 手動レビュー (Excel 上で並べて確認) を効率化する

依存: openpyxl (ScanToXlsxStarter の既存依存には含まれないため、
preview.py を使う場合は別途インストールする:
    pip install openpyxl
)
"""
from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:  # pragma: no cover
    raise ImportError(
        "preview.py は openpyxl に依存します。`pip install openpyxl` してください。"
    ) from e


# ------------------------------------------------------------
# ハイライト用スタイル
# ------------------------------------------------------------
FILL_HEADER = PatternFill("solid", fgColor="263238")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FILL_CHANGED = PatternFill("solid", fgColor="FFFDE7")     # 薄黄: 校正で変化
FILL_LEAD_CHANGED = PatternFill("solid", fgColor="FFE0B2") # オレンジ: 主要フィールド変化
FILL_WARN = PatternFill("solid", fgColor="FFF3E0")        # 薄オレンジ: 警告
FILL_ERROR = PatternFill("solid", fgColor="FFCDD2")       # 赤: エラー


def _label(row: Mapping[str, Any], label_fields: Sequence[str]) -> str:
    parts = [str(row.get(f, "")) for f in label_fields]
    parts = [p for p in parts if p]
    return "/".join(parts) or "(no-label)"


def write_comparison(
    output: Path,
    *,
    rows: Iterable[Mapping[str, Any]],
    fields: Sequence[str],
    before_key: str = "_before",
    after_key: str = "_after",
    label_fields: Sequence[str] = ("label",),
    emphasis_fields: Sequence[str] = (),
    warn_fn=None,
    column_widths: Mapping[str, int] | None = None,
    sheet_title: str = "preview",
) -> dict[str, int]:
    """元 (before) と 校正後 (after) を並べた比較 XLSX を書き出す。

    各 row は次の構造を持つ dict:
        {
            "label": "1/1",   # label_fields で取り出すキー
            "_before": {"lead": "...", "leadText": "..."},
            "_after":  {"lead": "...", "leadText": "..."},
            ...
        }

    列: [ラベル, フィールド1 (元), フィールド1 (校正後), フィールド2 (元), ..., 備考]

    Args:
        output: 出力 XLSX パス。
        rows: 上記構造の iterable。
        fields: 比較するフィールド名のリスト (例: ["lead", "leadText"])。
        before_key: row の中で「変更前」dict が入っているキー。
        after_key: row の中で「変更後」dict が入っているキー。
        label_fields: ラベル列を組み立てるキーのリスト。
        emphasis_fields: ハイライトを強めにする (= オレンジ) フィールド。
        warn_fn: row -> str|None の callable。返り値が str ならその文言を備考列に追加。
        column_widths: {列名: 幅} の辞書。未指定は自動。
        sheet_title: シート名。

    Returns:
        集計サマリ: {"rows": N, "changed": M, "warnings": W}
    """
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active
    ws.title = sheet_title

    # ヘッダー構築
    headers = ["ラベル"]
    for f in fields:
        headers.append(f"{f} (元)")
        headers.append(f"{f} (校正後)")
    headers.append("備考")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
    ws.freeze_panes = "A2"

    # 列幅
    default_widths = {"ラベル": 10, "備考": 24}
    for f in fields:
        default_widths[f"{f} (元)"] = 50
        default_widths[f"{f} (校正後)"] = 50
    if column_widths:
        default_widths.update(column_widths)
    for col, h in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = default_widths.get(h, 40)

    row_count = 0
    changed_rows = 0
    warn_rows = 0
    for r in rows:
        row_count += 1
        before = r.get(before_key) or {}
        after = r.get(after_key) or {}
        out_row = ws.max_row + 1
        ws.cell(row=out_row, column=1, value=_label(r, label_fields))

        notes: list[str] = []
        any_changed = False
        col_idx = 2
        for f in fields:
            v_before = before.get(f)
            v_after = after.get(f)
            c_b = ws.cell(row=out_row, column=col_idx, value=v_before)
            col_idx += 1
            c_a = ws.cell(row=out_row, column=col_idx, value=v_after)
            col_idx += 1
            for c in (c_b, c_a):
                c.alignment = Alignment(wrap_text=True, vertical="top")
            if (v_before or "") != (v_after or ""):
                any_changed = True
                c_a.fill = FILL_LEAD_CHANGED if f in emphasis_fields else FILL_CHANGED
                notes.append(f"{f}修正")

        if warn_fn is not None:
            w = warn_fn(r)
            if w:
                notes.append(w)
                warn_rows += 1
        if any_changed:
            changed_rows += 1

        note_cell = ws.cell(row=out_row, column=col_idx, value=", ".join(notes))
        if any("空" in n or "エラー" in n for n in notes):
            note_cell.fill = FILL_ERROR
        elif notes:
            note_cell.fill = FILL_WARN

    wb.save(output)
    return {"rows": row_count, "changed": changed_rows, "warnings": warn_rows}


def default_warn_fn(
    row: Mapping[str, Any],
    *,
    after_key: str = "_after",
    fields: Sequence[str] = ("lead", "leadText"),
    min_lengths: Mapping[str, int] | None = None,
    max_lengths: Mapping[str, int] | None = None,
) -> str | None:
    """よくある品質チェック: 空・短すぎ・長すぎを検出。"""
    after = row.get(after_key) or {}
    notes = []
    min_lengths = min_lengths or {}
    max_lengths = max_lengths or {}
    for f in fields:
        v = (after.get(f) or "").strip()
        if not v:
            notes.append(f"{f}空")
            continue
        lo = min_lengths.get(f)
        hi = max_lengths.get(f)
        if lo is not None and len(v) < lo:
            notes.append(f"{f}短い({len(v)}字)")
        if hi is not None and len(v) > hi:
            notes.append(f"{f}長い({len(v)}字)")
    return ", ".join(notes) or None
